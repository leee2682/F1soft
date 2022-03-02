import openpyxl
import shutil
from tqdm import tqdm
import json
import os

wb = openpyxl.Workbook()

wb.active.title = "log"

ws = wb.active

ws['A1'] = '변경전 filename'

ws['C1'] = '변경후 filename'

path = r'H:/73_20220128/'
move_path = r'H:/'

file_lst = os.listdir(path)

i = 1


def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print('Error: Creating directory. ' + directory)


for file in tqdm(file_lst):
    i += 1

    src = os.path.join(path, file)

    with open(src, "r", encoding="utf8") as f:
        data = json.load(f)

    json_split = file.split('_')
    json_name = '{}_{}_{}_{}_{}_{}_{}_{}'.format(json_split[0], json_split[1], data['description']['state'],
                                                 json_split[3], data['info']['city_id'], json_split[-3], json_split[-2],
                                                 json_split[-1])

    try:
        if int(json_split[1]) >= 15:
            createFolder(move_path + json_split[1] + '/' + data['description']['state'] + '/')

            dst = os.path.join(move_path + json_split[1] + '/' + data['description']['state'] + '/', json_name)

            ws.cell(i, 1, data['info']['filename'])

            file_name = json_name.split('.')[0]
            file_split = data['info']['filename']
            refile_name = '{}.{}'.format(file_name, file_split.split('.')[1])

            try:
                shutil.copy(src, dst)

                with open(dst, "r", encoding="utf8") as f:
                    data = json.load(f)

                data['info']['filename'] = refile_name
                with open(dst, 'w', encoding="utf8") as outfile:
                    json.dump(data, outfile, indent='\t', ensure_ascii=False)

                ws.cell(i, 3, data['info']['filename'])
            except:
                try:
                    createFolder(move_path + 'None/')

                    shutil.copy(src, os.path.join(move_path + 'None/', json_name))

                    data['info']['filename'] = refile_name

                    with open(os.path.join(move_path + 'None/', json_name), 'w', encoding="utf8") as outfile:
                        json.dump(data, outfile, indent='\t', ensure_ascii=False)
                except:
                    print(file)

    except:
        shutil.copy(src, os.path.join(move_path + 'None/', json_name))

wb.save('H:/119_log.xlsx')