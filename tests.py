from tqdm import tqdm
from openpyxl import load_workbook
import os

wb = load_workbook(filename = './119 최종데이터 제출 수량.xlsx', read_only = False, data_only = False)

wb.active.title = "Sheet1"
ws = wb.active

path = r'/share/CACHEDEV1_DATA/ProcessData_F1Soft/119/2. Label/'

clas = ['01', '02', '31', '32', '03', '04', '05', '06', '07', '08', '09', '10' ,'11' ,'12' ,'13' ,'14' ,'15' ,'16', '38','21', '23','24','25','26','27','28','29','30','36','37']

city = {'1':0, '2':0, '3':0, '4':0, '5':0, '6':0}

folder_lst = os.listdir(path)

i = 6
j = 15

for k in range(len(clas)):
    print(clas[k])
    fir = os.path.join(path, clas[k])
    for se in os.listdir(fir):
        print(se)
        se = os.path.join(fir, se)
        for th in os.listdir(se):
            print(th)
    #         city[th.split('_')[4]] += 1
        
    #     if (se.split('\\')[-1][:1] == '0'):
    #         ws.cell(i, 15, city['1'])
    #         ws.cell(i, 17, city['2'])
    #         ws.cell(i, 19, city['3'])
    #         ws.cell(i, 21, city['4'])
    #         ws.cell(i, 23, city['5'])
    #         ws.cell(i, 25, city['6'])
    #     else:
    #         ws.cell(i, 16, city['1'])
    #         ws.cell(i, 18, city['2'])
    #         ws.cell(i, 20, city['3'])
    #         ws.cell(i, 22, city['4'])
    #         ws.cell(i, 24, city['5'])
    #         ws.cell(i, 26, city['6'])
        
    #     print(city)
    #     city = {'1':0, '2':0, '3':0, '4':0, '5':0, '6':0}    
    # i += 1

# wb.save('C:/Users/admin/Desktop/test.xlsx')